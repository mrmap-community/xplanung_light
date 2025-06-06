from django.http import HttpResponse
from django.shortcuts import render
from xplanung_light.forms import RegistrationForm, BPlanCreateForm, BPlanUpdateForm
from django.shortcuts import redirect
from django.contrib.auth import login
from xplanung_light.models import AdministrativeOrganization
from django.contrib.gis.geos import GEOSGeometry
from openpyxl import Workbook, load_workbook
import requests
from django.views.generic import (ListView, CreateView, UpdateView, DeleteView)
from xplanung_light.models import AdministrativeOrganization, BPlan, BPlanSpezExterneReferenz
from django.urls import reverse_lazy
from leaflet.forms.widgets import LeafletWidget
from django_tables2 import SingleTableView
from xplanung_light.tables import BPlanTable
from django.views.generic import DetailView
from django.contrib.gis.db.models.functions import AsGML, Transform, Envelope
from django.contrib.gis.gdal import CoordTransform, SpatialReference
from django.contrib.gis.gdal import OGRGeometry
import uuid
import xml.etree.ElementTree as ET
from django.core.serializers import serialize
import json
from .filter import BPlanFilter
from django_filters.views import FilterView
from django.urls import reverse_lazy, reverse
from xplanung_light.helper.xplanung import XPlanung
from django.contrib import messages
from xplanung_light.forms import RegistrationForm, BPlanImportForm
from django.db.models import Subquery, OuterRef
from django.http import HttpResponse
import mapscript
from urllib.parse import parse_qs
from xplanung_light.helper.mapfile import MapfileGenerator
# for caching mapfiles ;-)
from django.core.cache import cache
from django.conf import settings
from xplanung_light.tables import BPlanTable, AdministrativeOrganizationPublishingTable, BPlanSpezExterneReferenzTable
from django.db.models import Count, F
from django.contrib.gis.db.models import Extent
from django.http import FileResponse
import os

"""
PROXIES = {
    'http_proxy': 'http://xxx:8080',
    'https_proxy': 'http://xxx:8080',
}
"""
PROXIES = None

def get_bplan_attachment(request, pk):
    try:
        #attachment = BPlanSpezExterneReferenz.objects.get(owned_by_user=request.user, pk=pk)
        attachment = BPlanSpezExterneReferenz.objects.get(pk=pk)
    except BPlanSpezExterneReferenz.DoesNotExist:
        attachment = None
    print(str(attachment))
    if attachment:
        if os.path.exists(attachment.attachment.file.name):
            response = FileResponse(attachment.attachment)
            return response
        else:
           return HttpResponse("File not found", status=404) 
    else:
        return HttpResponse("Object not found", status=404)

def ows(request, pk:int):
    orga = AdministrativeOrganization.objects.get(pk=pk)
    req =  mapscript.OWSRequest()
    """
    req.setParameter( 'SERVICE', 'WMS' )
    req.setParameter( 'VERSION', '1.1.0' )
    req.setParameter( 'REQUEST', 'GetCapabilities' )
    """
    #print(request.META['QUERY_STRING'])
    qs = parse_qs(request.META['QUERY_STRING'])
    for k, v in qs.items():
        #print(k)
        #print(v)
        req.setParameter(k, ','.join(v))
    #print(req)

    # test wfs http://127.0.0.1:8000/organization/1/ows/?REQUEST=GetFeature&VERSION=1.1.0&SERVICE=wfs&typename=BPlan.0723507001.12
    ## first variant - fast - 0.07 seconds
    
    #map = mapscript.mapObj( '/home/armin/devel/django/komserv2/test.map' )
    
    ## alternative approach - read from file into string and then from string with special path - also fast - 0.1 seconds
    
    #with open('/home/armin/devel/django/komserv2/test.map') as file:
        #map_file_string = file.read()
    #map = mapscript.msLoadMapFromString(map_file_string, '/home/armin/devel/django/komserv2/')
    
    ## next alternative - slowest - 1.1 seconds
    #mapfile = mappyfile.open("/home/armin/devel/django/komserv2/test.map")
    #map = mapscript.msLoadMapFromString(mappyfile.dumps(mapfile), '/home/armin/devel/django/komserv2/')

    ## next alternative - load from dynamically generated mapfile ;-)
    mapfile_generator = MapfileGenerator()
    metadata_uri = request.build_absolute_uri(reverse('bplan-export-iso19139', kwargs={"pk": 1000000}))
    # test to read mapfile from cache
    if cache.get("mapfile_" + orga.ags):
        cache.touch("mapfile_" + orga.ags, 10)
        mapfile = cache.get("mapfile_" + orga.ags)
    else:
        mapfile = mapfile_generator.generate_mapfile(pk, request.build_absolute_uri(reverse('ows', kwargs={"pk": pk})), metadata_uri)
        cache.set("mapfile_" + orga.ags, mapfile, 10)
    #print(mapfile)
    map = mapscript.msLoadMapFromString(mapfile, str(settings.BASE_DIR) + "/") 
    mapscript.msIO_installStdoutToBuffer()
    dispatch_status = map.OWSDispatch(req)

    if dispatch_status != mapscript.MS_SUCCESS:
        if dispatch_status == mapscript.MS_DONE:
            return HttpResponse("No valid OWS Request!")
        if dispatch_status == mapscript.MS_FAILURE:
            return HttpResponse("No valid OWS Request not successfully processed!")
    
    content_type = mapscript.msIO_stripStdoutBufferContentType()
    mapscript.msIO_stripStdoutBufferContentHeaders()
    result = mapscript.msIO_getStdoutBufferBytes()
    # [('Content-Type', 'application/vnd.ogc.wms_xml; charset=UTF-8'), ('Content-Length', '11385')]
    response_headers = [('Content-Type', content_type),
                        ('Content-Length', str(len(result)))]

    assert int(response_headers[1][1]) > 0

    http_response = HttpResponse(result)
    http_response.headers['Content-Type'] = content_type
    http_response.headers['Content-Length'] = str(len(result))
    return http_response

def bplan_import(request):
    if request.method == "POST":
        form = BPlanImportForm(request.POST, request.FILES)
        #print("bplan_import: form rendered")
        if form.is_valid():
            #print("bplan_import: form valid")
            # https://stackoverflow.com/questions/44722885/reading-inmemoryuploadedfile-twice
            # pointer muss auf Dateianfang gesetzt sein!
            request.FILES['file'].seek(0)
            xplanung = XPlanung(request.FILES["file"])
            # import xml file after prevalidation - check is done, if object already exists
            overwrite = form.cleaned_data['confirm']
            #print(overwrite)
            bplan_created = xplanung.import_bplan(overwrite=overwrite)
            if bplan_created == False:
                messages.error(request, 'Bebauungsplan ist schon vorhanden - bitte selektieren sie explizit \"Vorhandenen Plan überschreiben\"!')
                # extent form  with confirmation field!
                # https://amgcomputing.blogspot.com/2015/11/django-form-confirm-before-saving.html
                # reload form
                form = BPlanImportForm()
                return render(request, "xplanung_light/bplan_import.html", {"form": form})
            else:
                if overwrite:
                    messages.success(request, 'Bebauungsplan wurde erfolgreich aktualisiert!')
                else:
                    messages.success(request, 'Bebauungsplan wurde erfolgreich importiert!')
            #print("bplan_import: import done")
            return redirect(reverse('bplan-list'))
        else:
            print("bplan_import: form invalid")
    else:
        #print("bplan_import: no post")
        form = BPlanImportForm()
    return render(request, "xplanung_light/bplan_import.html", {"form": form})

def qualify_gml_geometry(gml_from_db:str):
    ET.register_namespace('gml','http://www.opengis.net/gml/3.2')
    root = ET.fromstring("<?xml version='1.0' encoding='UTF-8'?><snippet xmlns:gml='http://www.opengis.net/gml/3.2'>" + gml_from_db + "</snippet>")
    ns = {
        'gml': 'http://www.opengis.net/gml/3.2',
    }
    # print("<?xml version='1.0' encoding='UTF-8'?><snippet xmlns:gml='http://www.opengis.net/gml/3.2'>" + context['bplan'].geltungsbereich_gml_25832 + "</snippet>")
    # Test ob ein Polygon zurück kommt - damit wäre nur ein einziges Polygon im geometry Field
    polygons = root.findall('gml:Polygon', ns)
    # print(len(polygons))
    if len(polygons) == 0:
        # print("Kein Polygon auf oberer Ebene gefunden - es sind wahrscheinlich mehrere!")
        multi_polygon_element = root.find('gml:MultiSurface', ns)
        uuid_multisurface = uuid.uuid4()
        multi_polygon_element.set("gml:id", "GML_" + str(uuid_multisurface))
        # Füge gml_id Attribute hinzu - besser diese als Hash aus den Geometrien zu rechnen, oder in Zukunft generic_ids der Bereiche zu verwenden 
        polygons = root.findall('gml:MultiSurface/gml:surfaceMember/gml:Polygon', ns)
        for polygon in polygons:
            uuid_polygon = uuid.uuid4()
            polygon.set("gml:id", "GML_" + str(uuid_polygon))
        return ET.tostring(multi_polygon_element, encoding="utf-8", method="xml").decode('utf8')
    else:
        polygon_element = root.find('gml:Polygon', ns)
        #polygon_element.set("xmlns:gml", "http://www.opengis.net/gml/3.2")   
        uuid_polygon = uuid.uuid4()
        polygon_element.set("gml:id", "GML_" + str(uuid_polygon))
        # Ausgabe der Geometrie in ein XML-Snippet - erweitert um den MultiSurface/surfaceMember Rahmen
        ET.dump(polygon_element) 
        return '<gml:MultiSurface srsName="EPSG:25832"><gml:surfaceMember>' + ET.tostring(polygon_element, encoding="utf-8", method="xml").decode('utf8') + '</gml:surfaceMember></gml:MultiSurface>'


def get_geometry(type, ags):
    if type =='KR' or type =='KFS':
        base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:landkreise_rlp"
        param_dict = {'f': 'json', 'kreissch': ags[:3]}
    if type =='VG' or type =='VFG':
        base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:verbandsgemeinde_rlp"
        param_dict = {'f': 'json', 'vgnr': ags[:5]}
    if type == 'GE':
        base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:gemeinde_rlp"
        param_dict = {'f': 'json', 'ags': "*7" + ags}
    resp = requests.get(url=base_uri, params=param_dict, proxies=PROXIES)
    print(base_uri)
    print(str(param_dict))
    data = resp.json() 
    return str(data['features'][0]['geometry'])

def import_organisations():
    wb = load_workbook('Kommunalverwaltungen_01.01_2025.xlsm')
    # nuts-1 - bundeslandebene
    # nuts-2 - regierungsbezirke
    # nuts-3 - landkreisebene
    # lau-1 - verbandsgemeindeebene
    # lau-2 - gemeideebene

    table_all_admin_units = wb.worksheets[10]
    table_nuts_3_1 = wb.worksheets[5]
    table_nuts_3_2 = wb.worksheets[6]
    table_lau_1_1 = wb.worksheets[8]
    table_lau_1_2 = wb.worksheets[7]
    # table_lau_2 = wb.worksheets[10]

    count_landkreise = 0
    count_kreisfreie_staedte = 0
    count_verbandsgemeinden = 0
    count_verbandsfreie_gemeinden = 0
    count_gemeinden = 0
    # read landkreisebene
    landkreisebene = {}
    i = 0
    for row in table_nuts_3_1.iter_rows(values_only=True):
        i = i + 1
        if i > 2:
            if row[0] != None:
                landkreis = {}
                landkreis['kr'] = row[0]
                landkreis['vg'] = row[2]
                landkreis['ge'] = row[1]
                landkreis['name'] = row[4]
                landkreis['type'] = 'KR'
                landkreis['address'] = {}
                landkreis['address']['street'] = row[9]
                landkreis['address']['postcode'] = row[10]
                landkreis['address']['city'] = row[11]
                landkreis['address']['phone'] = str(row[12]) + '/' + str(row[13])
                landkreis['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                landkreis['address']['email'] = str(row[15])
                landkreis['address']['homepage'] = "https://" + str(row[16])
                landkreisebene[row[0] + row[2] + row[1]] = landkreis
                count_landkreise = count_landkreise + 1
    i = 0           
    for row in table_nuts_3_2.iter_rows(values_only=True):
        i = i + 1
        if i > 2:
            if row[0] != None:
                kreisfreie_stadt = {}
                kreisfreie_stadt['kr'] = row[0]
                kreisfreie_stadt['vg'] = row[2]
                kreisfreie_stadt['ge'] = row[1]
                kreisfreie_stadt['name'] = row[4]
                kreisfreie_stadt['type'] = 'KFS'
                kreisfreie_stadt['address'] = {}
                kreisfreie_stadt['address']['street'] = row[9]
                kreisfreie_stadt['address']['postcode'] = row[10]
                kreisfreie_stadt['address']['city'] = row[11]
                kreisfreie_stadt['address']['phone'] = str(row[12]) + '/' + str(row[13])
                kreisfreie_stadt['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                kreisfreie_stadt['address']['email'] = str(row[15])
                kreisfreie_stadt['address']['homepage'] = "https://" + str(row[16])
                landkreisebene[row[0] + row[2] + row[1]] = kreisfreie_stadt
                count_kreisfreie_staedte = count_kreisfreie_staedte + 1
    # read verbandsgemeindeebene
    verbandsgemeindeebene = {}
    i = 0
    for row in table_lau_1_1.iter_rows(values_only=True):
        i = i + 1
        if i > 2:
            if row[0] != None:
                vg = {}
                vg['kr'] = row[0]
                vg['vg'] = row[2]
                vg['ge'] = row[1]
                vg['name'] = row[4]
                vg['type'] = 'VG'
                vg['address'] = {}
                vg['address']['street'] = row[9]
                vg['address']['postcode'] = row[10]
                vg['address']['city'] = row[11]
                vg['address']['phone'] = str(row[12]) + '/' + str(row[13])
                vg['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                vg['address']['email'] = str(row[15])
                vg['address']['homepage'] = "https://" + str(row[16])
                verbandsgemeindeebene[row[0] + row[2] + row[1]] = vg
                count_verbandsgemeinden = count_verbandsgemeinden + 1
    i = 0
    for row in table_lau_1_2.iter_rows(values_only=True):
        i = i + 1
        if i > 2:
            if row[0] != None:
                vg = {}
                vg['kr'] = row[0]
                vg['vg'] = row[2]
                vg['ge'] = row[1]
                vg['name'] = row[4]
                vg['type'] = 'VFG'
                vg['address'] = {}
                vg['address']['street'] = row[9]
                vg['address']['postcode'] = row[10]
                vg['address']['city'] = row[11]
                vg['address']['phone'] = str(row[12]) + '/' + str(row[13])
                vg['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                vg['address']['email'] = str(row[15])
                vg['address']['homepage'] = "https://" + str(row[16])
                verbandsgemeindeebene[row[0] + row[2] + row[1]] = vg
                count_verbandsfreie_gemeinden = count_verbandsfreie_gemeinden + 1
    #print(json.dumps(landkreise))
    all_admin_units = {}
    i = 0
    for row in table_all_admin_units.iter_rows(values_only=True):
        i = i + 1
        if i > 2:
            if row[0] != None:
                admin_unit = {}
                admin_unit['kr'] = row[0]
                admin_unit['vg'] = row[1]
                admin_unit['ge'] = row[2]
                admin_unit['name'] = row[3]
                print(admin_unit['name'])
                admin_unit['plz'] = row[4]
                admin_unit['type'] = 'GE'
                if row[1] == '00' and row[2] == '000':
                    admin_unit['type'] = landkreisebene[row[0] + row[1] + row[2]]['type']
                    admin_unit['address'] = landkreisebene[row[0] + row[1] + row[2]]['address']
                if row[1] != '00' and row[2] == '000':
                    admin_unit['type'] = verbandsgemeindeebene[row[0] + row[1] + row[2]]['type']
                    admin_unit['address'] = verbandsgemeindeebene[row[0] + row[1] + row[2]]['address']
                admin_unit['geometry'] = get_geometry(admin_unit['type'], str(row[0]) + str(row[1]) + str(row[2]))
                all_admin_units[str(row[0]) + str(row[1]) +str(row[2])] = admin_unit   
                #save object to database
                obj, created = AdministrativeOrganization.objects.update_or_create(
                    ks=admin_unit['kr'],
                    vs=admin_unit['vg'],
                    gs=admin_unit['ge'],
                    defaults={
                        "ks": admin_unit['kr'],
                        "vs": admin_unit['vg'],
                        "gs": admin_unit['ge'],
                        "name": admin_unit['name'],
                        "type": admin_unit['type'],
                        "geometry": GEOSGeometry(admin_unit['geometry'])
                    },
                )
                """
                administration = AdministrativeOrganization()
                administration.ks = admin_unit['kr']
                administration.vs = admin_unit['vg']
                administration.gs = admin_unit['ge']
                administration.name = admin_unit['name']
                administration.type = admin_unit['type']
                administration.geometry = GEOSGeometry(admin_unit['geometry'])
                administration.save()
                """

    print("Landkreise:" + str(count_landkreise))
    print("Kreisfreie Städte:" + str(count_kreisfreie_staedte))
    print("Verbandsgemeinden:" + str(count_verbandsgemeinden))
    print("Verbandsfreie Gemeinden:" + str(count_verbandsfreie_gemeinden))
    print(i)

def home(request):
    return render(request, "xplanung_light/home.html")
    
def about(request):
    return render(request, "xplanung_light/about.html")

# https://dev.to/balt1794/registration-page-using-usercreationform-django-part-1-21j7
def register(request):
    if request.method != 'POST':
        form = RegistrationForm()
    else:
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            user = form.save()
            login(request, user)
            return redirect('home')
        else:
            print('form is invalid')
    context = {'form': form}
    return render(request, 'registration/register.html', context)


class BPlanCreateView(CreateView):
    form_class = BPlanCreateForm
    model = BPlan
    # copy fields to form class - cause form class will handle the form now!
    #fields = ["name", "nummer", "geltungsbereich", "gemeinde", "planart", "inkrafttretens_datum", "staedtebaulicher_vetrag"]
    success_url = reverse_lazy("bplan-list") 

    def get_form(self, form_class=None):
        form = super().get_form(self.form_class)
        form.fields['gemeinde'].queryset = form.fields['gemeinde'].queryset.annotate(bbox=(Extent("geometry"))).only("pk", "name", "type")
        form.fields['geltungsbereich'].widget = LeafletWidget(attrs={'geom_type': 'MultiPolygon', 'map_height': '400px', 'map_width': '90%','MINIMAP': True})
        return form
    

class BPlanUpdateView(UpdateView):
    form_class = BPlanUpdateForm
    model = BPlan
    success_url = reverse_lazy("bplan-list") 

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['gemeinde'].queryset = form.fields['gemeinde'].queryset.annotate(bbox=(Extent("geometry"))).only("pk", "name", "type")
        form.fields['geltungsbereich'].widget = LeafletWidget(attrs={'geom_type': 'MultiPolygon', 'map_height': '400px', 'map_width': '90%','MINIMAP': True})
        return form


class BPlanDeleteView(DeleteView):
    model = BPlan

    def get_success_url(self):
        return reverse_lazy("bplan-list")


class BPlanListView(FilterView, SingleTableView):
    model = BPlan
    table_class = BPlanTable
    template_name = 'xplanung_light/bplan_list.html'
    success_url = reverse_lazy("bplan-list") 
    filterset_class = BPlanFilter

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # TODO: Anstatt object_list.data vlt. table.data? ... - dann haben wir mehr Einfluss auf die Darstellung im Leaflet Client
        context["markers"] = json.loads(
            serialize("geojson", context['table'].page.object_list.data, geometry_field='geltungsbereich')
        )
        #print(context["markers"])
        return context

    def get_queryset(self):
        #qs = super().get_queryset()
        #https://github.com/jazzband/django-simple-history/issues/407
        # https://stackoverflow.com/questions/43364451/how-to-get-the-last-changed-object-in-django-simple-history
        
        qs = BPlan.objects.select_related('gemeinde').annotate(last_changed=Subquery(
            BPlan.history.filter(id=OuterRef("pk")).order_by('-history_date').values('history_date')[:1]
        )).order_by('-last_changed').annotate(bbox=Envelope("geltungsbereich")).prefetch_related('attachments')

        self.filter_set = BPlanFilter(self.request.GET, queryset=qs)
        return self.filter_set.qs


class BPlanDetailView(DetailView):
    model = BPlan


class BPlanDetailXmlRasterView(BPlanDetailView):  

    def get_queryset(self):
        # Erweiterung der auszulesenden Objekte um eine transformierte Geomtrie im Format GML 3
        queryset = super().get_queryset().annotate(geltungsbereich_gml_25832=AsGML(Transform("geltungsbereich", 25832), version=3)).annotate(geltungsbereich_gml_4326=AsGML("geltungsbereich", version=3))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Um einen XPlanung-konformen Auszug zu bekommen, werden gml_id(s) verwendet.
        # Es handelt sich um uuids, die noch das Prefix "GML_" bekommen. Grundsätzlich sollten die 
        # aus den Daten in der DB stammen und dort vergeben werden. 
        # Im ersten Schritt synthetisieren wir sie einfach ;-)
        context['auszug_uuid'] = "GML_" + str(uuid.uuid4())
        context['bplan_uuid'] = "GML_" + str(uuid.uuid4())
        # Irgendwie gibt es keine django model function um direkt den Extent der Geometrie zu erhalten. Daher nutzen wir hier gdal
        # und Transformieren die Daten erneut im RAM
        # Definition der Transformation (Daten sind immer in WGS 84 - 4326)
        ct = CoordTransform(SpatialReference(4326, srs_type='epsg'), SpatialReference(25832, srs_type='epsg'))
        # OGRGeoemtry Objekt erstellen
        ogr_geom = OGRGeometry(str(context['bplan'].geltungsbereich), srs=4326)
        context['wgs84_extent'] = ogr_geom.extent
        # Transformation nach EPSG:25832
        ogr_geom.transform(ct)
        # Speichern des Extents in den Context
        context['extent'] = ogr_geom.extent
        # Ausgabe der GML Variante zu Testzwecken 
        # print(context['bplan'].geltungsbereich_gml_25832)
        # Da die GML Daten nicht alle Attribute beinhalten, die XPlanung fordert, müssen wir sie anpassen, bzw. umschreiben
        # Hierzu nutzen wir die Funktion qualify_gml_geometry
        context['multisurface_geometry_25832'] = qualify_gml_geometry(context['bplan'].geltungsbereich_gml_25832)
        context['multisurface_geometry_4326'] = qualify_gml_geometry(context['bplan'].geltungsbereich_gml_4326)

        relative_url = reverse('bplan-export-xplan-raster-6', kwargs={'pk': context['bplan'].id})
        context['iso19139_url']= self.request.build_absolute_uri(relative_url)
        return context

    def dispatch(self, *args, **kwargs):
        response = super().dispatch(*args, **kwargs)
        response['Content-type'] = "application/xml"  # set header
        return response


class AdministrativeOrganizationPublishingListView(SingleTableView):
    model = AdministrativeOrganization
    table_class = AdministrativeOrganizationPublishingTable
    template_name = 'xplanung_light/orga_publishing_list.html'
    success_url = reverse_lazy("orga-publishing-list") 

    def get_queryset(self):
        qs = AdministrativeOrganization.objects.filter(bplan__isnull=False).distinct().annotate(num_bplan=Count('bplan'))
        return qs
    

class BPlanSpezExterneReferenzCreateView(CreateView):
    model = BPlanSpezExterneReferenz
    fields = ["typ", "name", "attachment"]
    
    def get_context_data(self, **kwargs):
        bplanid = self.kwargs['bplanid']
        context = super().get_context_data(**kwargs)
        context['bplan'] = BPlan.objects.get(pk=bplanid)
        return context

    # reduce choices for invoice to own invoices    
    # https://stackoverflow.com/questions/48089590/limiting-choices-in-foreign-key-dropdown-in-django-using-generic-views-createv
    def get_form(self, form_class=None):
        form = super().get_form(form_class=None)
        #form.fields['bplan'].queryset = form.fields['bplan'].queryset.filter(owned_by_user=self.request.user.id)
        #https://django-bootstrap-datepicker-plus.readthedocs.io/en/latest/Walkthrough.html
        #form.fields['issue_date'].widget = DatePickerInput()
        #form.fields['due_date'].widget = DatePickerInput()
        #form.fields['actual_delivery_date'].widget = DatePickerInput()
        return form
    
    def get_form_kwargs(self):
        form = super().get_form_kwargs()
        bplanid = self.kwargs['bplanid']
        
        form['initial'].update({'bplan': BPlan.objects.get(pk=bplanid)})
        #form['initial'].update({'owned_by_user': self.request.user})
        return form
        #return super().get_form_kwargs()

    def form_valid(self, form):
        bplanid = self.kwargs['bplanid']
        form.instance.bplan = BPlan.objects.get(pk=bplanid)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("bplanattachment-list", kwargs={'bplanid': self.kwargs['bplanid']})   
    

class BPlanSpezExterneReferenzListView(SingleTableView):
    model = BPlanSpezExterneReferenz
    table_class = BPlanSpezExterneReferenzTable
    template_name = 'xplanung_light/bplanspezexternereferenz_list.html'
    success_url = reverse_lazy("bplanattachment-list") 

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['bplanid'] = self.kwargs['bplanid']
        context['bplan'] = BPlan.objects.get(pk=self.kwargs['bplanid'])
        return context
    
    def get_queryset(self):
        # reduce queryset to those invoicelines which came from the invoice
        bplanid = self.kwargs['bplanid']
        if bplanid:
            return self.model.objects.filter(
            bplan=BPlan.objects.get(pk=bplanid)
        )#.order_by('-created')
        else:
            return self.model.objects#.order_by('-created')
        

class BPlanSpezExterneReferenzUpdateView(UpdateView):
    model = BPlanSpezExterneReferenz
    fields = ["typ", "name", "attachment"]

    def get_form(self, form_class=None):
        form = super().get_form(form_class=None)
        #form.fields['bplan'].queryset = form.fields['bplan'].queryset.filter(owned_by_user=self.request.user.id)
        return form 
    
    def form_valid(self, form):
        bplanid = self.kwargs['bplanid']
        form.instance.bplan = BPlan.objects.get(pk=bplanid)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("bplanattachment-list", kwargs={'bplanid': self.kwargs['bplanid']})


class BPlanSpezExterneReferenzDeleteView(DeleteView):
    model = BPlanSpezExterneReferenz

    def get_success_url(self):
        return reverse_lazy("bplanattachment-list", kwargs={'bplanid': self.kwargs['bplanid']})
    

