from pathlib import Path
from typing import Any, Optional

import requests
from django.contrib.gis.geos import GEOSGeometry
from django.core.management.base import BaseCommand, CommandError, CommandParser
from openpyxl import load_workbook

from xplanung_light.models import AdministrativeOrganization


"""
PROXIES = {
    'http': 'http://xxx:8080',
    'https': 'http://xxx:8080',
}
"""
PROXIES = None


class Command(BaseCommand):
    requires_migrations_checks = True

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument("file", type=Path, help="MS Excel file")

    def handle(self, *args: Any, **options: Any) -> Optional[str]:
        self.import_organisations(options["file"])

    def get_geometry(self, type: str, ags: str) -> str:
        if type =='KR' or type =='KFS':
            base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:landkreise_rlp"
            param_dict = {'f': 'json', 'kreissch': ags[:3]}
        elif type =='VG' or type =='VFG':
            base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:verbandsgemeinde_rlp"
            param_dict = {'f': 'json', 'vgnr': ags[:5]}
        elif type == 'GE':
            base_uri = "https://www.geoportal.rlp.de/spatial-objects/314/collections/vermkv:gemeinde_rlp"
            param_dict = {'f': 'json', 'ags': "*7" + ags}
        else:
            raise CommandError(f"Unbekannter Typ: {type}")
        resp = requests.get(url=base_uri, params=param_dict, proxies=PROXIES)
        # resp = requests.get(url=base_uri, params=param_dict)
        self.stdout.write(base_uri)
        self.stdout.write(str(param_dict))
        data = resp.json()
        return str(data['features'][0]['geometry'])

    def import_organisations(self, excel_file: Path):
        wb = load_workbook(excel_file)
        # nuts-1 - bundeslandebene
        # nuts-2 - regierungsbezirke
        # nuts-3 - landkreisebene
        # lau-1 - verbandsgemeindeebene
        # lau-2 - gemeideebene

        table_all_admin_units = wb.worksheets[10]
        table_nuts_3_1 = wb.worksheets[5]
        table_nuts_3_2 = wb.worksheets[6]
        table_lau_1_1 = wb.worksheets[8]
        table_lau_1_2 = wb.worksheets[7]
        # table_lau_2 = wb.worksheets[10]

        count_landkreise = 0
        count_kreisfreie_staedte = 0
        count_verbandsgemeinden = 0
        count_verbandsfreie_gemeinden = 0
        count_gemeinden = 0
        # read landkreisebene
        landkreisebene = {}
        i = 0
        for row in table_nuts_3_1.iter_rows(values_only=True):
            i = i + 1
            if i > 2:
                if row[0] != None:
                    landkreis = {}
                    landkreis['kr'] = row[0]
                    landkreis['vg'] = row[2]
                    landkreis['ge'] = row[1]
                    landkreis['name'] = row[4]
                    landkreis['type'] = 'KR'
                    landkreis['address'] = {}
                    landkreis['address']['street'] = row[9]
                    landkreis['address']['postcode'] = row[10]
                    landkreis['address']['city'] = row[11]
                    landkreis['address']['phone'] = str(row[12]) + '/' + str(row[13])
                    landkreis['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                    landkreis['address']['email'] = str(row[15])
                    landkreis['address']['homepage'] = "https://" + str(row[16])
                    landkreisebene[row[0] + row[2] + row[1]] = landkreis
                    count_landkreise = count_landkreise + 1
        i = 0           
        for row in table_nuts_3_2.iter_rows(values_only=True):
            i = i + 1
            if i > 2:
                if row[0] != None:
                    kreisfreie_stadt = {}
                    kreisfreie_stadt['kr'] = row[0]
                    kreisfreie_stadt['vg'] = row[2]
                    kreisfreie_stadt['ge'] = row[1]
                    kreisfreie_stadt['name'] = row[4]
                    kreisfreie_stadt['type'] = 'KFS'
                    kreisfreie_stadt['address'] = {}
                    kreisfreie_stadt['address']['street'] = row[9]
                    kreisfreie_stadt['address']['postcode'] = row[10]
                    kreisfreie_stadt['address']['city'] = row[11]
                    kreisfreie_stadt['address']['phone'] = str(row[12]) + '/' + str(row[13])
                    kreisfreie_stadt['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                    kreisfreie_stadt['address']['email'] = str(row[15])
                    kreisfreie_stadt['address']['homepage'] = "https://" + str(row[16])
                    landkreisebene[row[0] + row[2] + row[1]] = kreisfreie_stadt
                    count_kreisfreie_staedte = count_kreisfreie_staedte + 1
        # read verbandsgemeindeebene
        verbandsgemeindeebene = {}
        i = 0
        for row in table_lau_1_1.iter_rows(values_only=True):
            i = i + 1
            if i > 2:
                if row[0] != None:
                    vg = {}
                    vg['kr'] = row[0]
                    vg['vg'] = row[2]
                    vg['ge'] = row[1]
                    vg['name'] = row[4]
                    vg['type'] = 'VG'
                    vg['address'] = {}
                    vg['address']['street'] = row[9]
                    vg['address']['postcode'] = row[10]
                    vg['address']['city'] = row[11]
                    vg['address']['phone'] = str(row[12]) + '/' + str(row[13])
                    vg['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                    vg['address']['email'] = str(row[15])
                    vg['address']['homepage'] = "https://" + str(row[16])
                    verbandsgemeindeebene[row[0] + row[2] + row[1]] = vg
                    count_verbandsgemeinden = count_verbandsgemeinden + 1
        i = 0
        for row in table_lau_1_2.iter_rows(values_only=True):
            i = i + 1
            if i > 2:
                if row[0] != None:
                    vg = {}
                    vg['kr'] = row[0]
                    vg['vg'] = row[2]
                    vg['ge'] = row[1]
                    vg['name'] = row[4]
                    vg['type'] = 'VFG'
                    vg['address'] = {}
                    vg['address']['street'] = row[9]
                    vg['address']['postcode'] = row[10]
                    vg['address']['city'] = row[11]
                    vg['address']['phone'] = str(row[12]) + '/' + str(row[13])
                    vg['address']['facsimile'] = str(row[12]) + '/' + str(row[14])
                    vg['address']['email'] = str(row[15])
                    vg['address']['homepage'] = "https://" + str(row[16])
                    verbandsgemeindeebene[row[0] + row[2] + row[1]] = vg
                    count_verbandsfreie_gemeinden = count_verbandsfreie_gemeinden + 1
        # self.stdout.write(json.dumps(landkreise))
        all_admin_units = {}
        i = 0
        for row in table_all_admin_units.iter_rows(values_only=True):
            i = i + 1
            if i > 2:
                if row[0] != None:
                    admin_unit = {}
                    admin_unit['kr'] = row[0]
                    admin_unit['vg'] = row[1]
                    admin_unit['ge'] = row[2]
                    admin_unit['name'] = row[3]
                    self.stdout.write(admin_unit['name'])
                    admin_unit['plz'] = row[4]
                    admin_unit['type'] = 'GE'
                    if row[1] == '00' and row[2] == '000':
                        admin_unit['type'] = landkreisebene[row[0] + row[1] + row[2]]['type']
                        admin_unit['address'] = landkreisebene[row[0] + row[1] + row[2]]['address']
                    if row[1] != '00' and row[2] == '000':
                        admin_unit['type'] = verbandsgemeindeebene[row[0] + row[1] + row[2]]['type']
                        admin_unit['address'] = verbandsgemeindeebene[row[0] + row[1] + row[2]]['address']
                    admin_unit['geometry'] = self.get_geometry(admin_unit['type'], str(row[0]) + str(row[1]) + str(row[2]))
                    all_admin_units[str(row[0]) + str(row[1]) +str(row[2])] = admin_unit   
                    # save object to database
                    obj, created = AdministrativeOrganization.objects.update_or_create(
                        ks=admin_unit['kr'],
                        vs=admin_unit['vg'],
                        gs=admin_unit['ge'],
                        defaults={
                            "ks": admin_unit['kr'],
                            "vs": admin_unit['vg'],
                            "gs": admin_unit['ge'],
                            "name": admin_unit['name'],
                            "type": admin_unit['type'],
                            "geometry": GEOSGeometry(admin_unit['geometry'])
                        },
                    )
                    """
                    administration = AdministrativeOrganization()
                    administration.ks = admin_unit['kr']
                    administration.vs = admin_unit['vg']
                    administration.gs = admin_unit['ge']
                    administration.name = admin_unit['name']
                    administration.type = admin_unit['type']
                    administration.geometry = GEOSGeometry(admin_unit['geometry'])
                    administration.save()
                    """

        self.stdout.write("Landkreise:" + str(count_landkreise))
        self.stdout.write("Kreisfreie Städte:" + str(count_kreisfreie_staedte))
        self.stdout.write("Verbandsgemeinden:" + str(count_verbandsgemeinden))
        self.stdout.write("Verbandsfreie Gemeinden:" + str(count_verbandsfreie_gemeinden))
        self.stdout.write(str(i))